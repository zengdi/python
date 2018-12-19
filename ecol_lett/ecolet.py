import requests
from lxml import etree
import pandas as pd
import time
from openpyxl import Workbook

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36'
headers = {"User-Agent": user_agent}

html = requests.get('https://onlinelibrary.wiley.com/loi/14610248', headers=headers).content
selector = etree.HTML(html)
volumes = selector.xpath('//ul[@class="rlist loi__list"]//li/a/@href')
issues = []
base_http = 'https://onlinelibrary.wiley.com'
doi_http = 'https://doi.org'

for volume in volumes:
    http_volume = base_http + volume
    # print(http_volume)
    html_volume = requests.get(http_volume, headers=headers).content
    volume_selector = etree.HTML(html_volume)
    issue_temp = volume_selector.xpath('//ul[@class="rlist loi__issues"]//h4/a/@href')
    issues.extend(issue_temp)
    # time.sleep(1)
    break

dois = []
for issue in issues:
    http_issue = base_http + issue
    # print(http_issue)
    html_issue = requests.get(http_issue, headers=headers).content
    issue_selector = etree.HTML(html_issue)
    #date = issue_selector.xpath("//div[@class='cover-image__date']/span")
    doi = issue_selector.xpath("//div[@class='issue-item']/a[@class='issue-item__title']/@href")
    # print(heading)
    dois.extend(doi)
    # time.sleep(0.2)
    # break
# print(dois)
papers = {}

i = 0
j = 0
fails = []
for doi in dois:
    http_doi = doi_http + doi.replace('/doi', '')
    print(http_doi)
    i += 1
    if not requests.get(http_doi, headers=headers).status_code == 200:
        continue
    try:
        html_doi = requests.get(http_doi, headers=headers).content
        paper_selector = etree.HTML(html_doi)
        title = paper_selector.xpath("//div[@class='article-citation']/div[@class='citation']/h2/text()")
        authors = paper_selector.xpath("//div[@class='citation']/div/div/a/span/text()")
        address = paper_selector.xpath("//div[@class='citation']//div[1]/div[@id='a1']/p/text()")
        date = paper_selector.xpath("//div[@class='epub-section']/span[@class='epub-date']/text()")
        if len(address) > 0:
            if 'Corresponding Author' in address[0]:
                # print(address[2:])
                address = address[2:]
        # print(address)

        # keywords = paper_selector.xpath("//section[@class='keywords']/ul[@class='rlist rlist--inline']/li/a[@class='badge-type']")
        # received_date = paper_selector.xpath("//section[@class='publication-history']/ul[@class='rlist']/li[5]")
        # accepted_date = paper_selector.xpath("//section[@class='publication-history']/ul[@class='rlist']/li[3]")
        paper_info = {'date':','.join(date),
                      'author': ','.join(authors),
                      'addresss': ','.join(address),
                      # 'keywords':','.join(keywords),
                      # 'rece':receiprint('Finished:', i/len(dois))ved_date,
                      # 'accept':accepted_date
                      }
        papers[title[0]] = paper_info
        time.sleep(0.5)
        # break
    except Exception as e:
        print('DOI:', doi)
        print('Error', e)
        j += 1
        fails.append(doi)
    finally:
        print(papers, '-----')
        data = pd.DataFrame.from_dict(papers, orient='index')
        data.to_csv('data/paper.csv', mode='a+', header=False, encoding='utf-8')
        print('Finished:', i / len(dois))
        print('Failed:', j / len(dois))
        papers.clear()

'''
Title:
//div[@class='article-citation']/div[@class='citation']/h2/text()

Authors:
//div[@class='citation']/div/div/a/span/text()

First Author's address:
//div[@class='citation']//div[1]/div[@id='a1']/p/text()

keywords of each paper:
//section[@class='keywords']/ul[@class='rlist rlist--inline']/li/a[@class='badge-type']

Manuscript received:
//section[@class='publication-history']/ul[@class='rlist']/li[5]

Manuscript accepted:
//section[@class='publication-history']/ul[@class='rlist']/li[3]
'''

# wb = Workbook()
# ws = wb.active
#
# i = 2
# ws.cell(1,1).value = 'Title'
# for heading in headings:
#     ws.cell(i, 1).value = heading
#     i += 1
# wb.save(r'data/paper.xlsx')
